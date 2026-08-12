"""Connector-local document processing and semantic knowledge boundary."""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import math
import re
import time
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, BinaryIO, Protocol
from uuid import UUID, uuid5

import httpx
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import Settings
from app.infrastructure.database.models.document import DocumentModel
from app.infrastructure.database.repositories.operations import SqlAlchemyOperationsRepository

logger = logging.getLogger(__name__)
KNOWLEDGE_SCHEMA_VERSION = "1"
MINIMUM_EVIDENCE_SCORE = 0.60
MINIMUM_LEXICAL_MATCHES = 2
_LEXICAL_WEIGHT = 0.80
_VECTOR_WEIGHT = 0.20
_SEARCH_CANDIDATE_LIMIT = 100
_STOP_WORDS = frozenset(
    {
        "a",
        "after",
        "an",
        "and",
        "are",
        "for",
        "from",
        "how",
        "i",
        "in",
        "is",
        "it",
        "of",
        "on",
        "should",
        "the",
        "this",
        "to",
        "what",
        "when",
        "where",
        "which",
        "with",
    }
)


class KnowledgeUnavailableError(RuntimeError):
    pass


class KnowledgeIdentityError(ValueError):
    pass


@dataclass(frozen=True)
class KnowledgeSection:
    text: str
    page_number: int | None = None
    sheet_name: str | None = None
    row_start: int | None = None
    row_end: int | None = None
    section_title: str | None = None


@dataclass(frozen=True)
class KnowledgeChunk:
    chunk_id: UUID
    document_id: UUID
    content: str
    vector: list[float]
    metadata: dict[str, Any]


@dataclass(frozen=True)
class KnowledgeSearchResult:
    document_id: UUID
    chunk_id: UUID
    content: str
    score: float
    source: str
    metadata: dict[str, Any]


@dataclass(frozen=True)
class KnowledgeStats:
    document_count: int
    indexed_chunk_count: int
    pending_count: int
    failed_count: int
    last_index_activity: datetime | None
    qdrant_reachable: bool
    collection_available: bool
    documents_stored: bool
    vectors_stored: bool
    search_operational: bool
    last_search_success: bool | None
    last_search_at: datetime | None
    engine_version: str | None
    statistics_readable: bool


@dataclass(frozen=True)
class KnowledgeStoreHealth:
    qdrant_reachable: bool
    collection_available: bool
    engine_version: str | None = None
    statistics_readable: bool = True


class KnowledgeStore(Protocol):
    async def initialize(self, dimension: int) -> None: ...
    async def health_check(self) -> bool: ...
    async def health_status(self) -> KnowledgeStoreHealth: ...
    async def upsert_chunks(self, tenant_id: UUID, chunks: list[KnowledgeChunk]) -> None: ...
    async def search(
        self,
        tenant_id: UUID,
        vector: list[float],
        top_k: int,
        filters: dict[str, str] | None = None,
    ) -> list[KnowledgeSearchResult]: ...
    async def delete_document(self, tenant_id: UUID, document_id: UUID) -> None: ...
    async def delete_tenant_knowledge(self, tenant_id: UUID) -> None: ...
    async def count(self, tenant_id: UUID) -> int: ...


class ContentSanitizer:
    """Small, explicit pre-persistence secret filter; not a general DLP product."""

    _patterns = (
        re.compile(
            r"-----BEGIN (?:RSA |EC |OPENSSH |DSA )?PRIVATE KEY-----.*?"
            r"-----END (?:RSA |EC |OPENSSH |DSA )?PRIVATE KEY-----",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(r"(?im)^\s*Authorization\s*:\s*[^\r\n]+$"),
        re.compile(r"(?i)\bBearer\s+[A-Za-z0-9._~+/=-]{12,}"),
        re.compile(r"\bAKIA[0-9A-Z]{16}\b"),
        re.compile(r"\bgh[pousr]_[A-Za-z0-9]{20,}\b"),
        re.compile(
            r"(?im)^\s*(?:password|passwd|client_secret|api[_-]?token|access[_-]?token|"
            r"connector[_-]?secret|registration[_-]?token)\s*[:=]\s*[^\r\n]+$"
        ),
    )

    def sanitize(self, value: str) -> str:
        normalized = value.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
        for pattern in self._patterns:
            normalized = pattern.sub("[REDACTED]", normalized)
        return "\n".join(line.rstrip() for line in normalized.splitlines()).strip()


class LocalHashEmbeddingService:
    """Network-free feature hashing so customer text never leaves the connector to embed."""

    name = "peka-local-feature-hash"
    model = "token-bigram-v1"

    def __init__(self, dimension: int) -> None:
        self.dimension = dimension

    def embed(self, texts: list[str]) -> list[list[float]]:
        vectors: list[list[float]] = []
        for text in texts:
            vector = [0.0] * self.dimension
            tokens = re.findall(r"[\w.-]+", text.casefold())
            features = [
                *tokens,
                *(f"{left}_{right}" for left, right in zip(tokens, tokens[1:], strict=False)),
            ]
            for feature in features:
                digest = hashlib.blake2b(feature.encode("utf-8"), digest_size=8).digest()
                bucket = int.from_bytes(digest[:4], "big") % self.dimension
                vector[bucket] += 1.0 if digest[4] & 1 else -1.0
            norm = math.sqrt(sum(item * item for item in vector)) or 1.0
            vectors.append([item / norm for item in vector])
        return vectors


def _normalize_relevance_token(token: str) -> str:
    """Normalize common inflections without introducing a language/model dependency."""
    if len(token) > 5 and token.endswith("ing"):
        return token[:-3]
    if len(token) > 4 and token.endswith("ied"):
        return token[:-3] + "y"
    if len(token) > 4 and token.endswith("ed"):
        return token[:-2]
    if len(token) > 4 and token.endswith("ly"):
        return token[:-2]
    if len(token) > 3 and token.endswith("s") and not token.endswith("ss"):
        return token[:-1]
    return token


def _relevance_terms(text: str) -> set[str]:
    return {
        _normalize_relevance_token(token)
        for token in re.findall(r"[a-z0-9]+", text.casefold())
        if len(token) > 1 and token not in _STOP_WORDS
    }


def _lexical_relevance(query: str, content: str) -> tuple[float, int]:
    query_terms = _relevance_terms(query)
    if not query_terms:
        return 0.0, 0
    matched_terms = len(query_terms & _relevance_terms(content))
    return matched_terms / len(query_terms), matched_terms


class QdrantKnowledgeStore:
    """The only connector component that knows the Qdrant REST contract."""

    _indexed_fields = ("tenant_id", "document_id", "source_id", "content_hash", "schema_version")

    def __init__(self, url: str, collection: str, timeout: float = 10) -> None:
        self.url = url.rstrip("/")
        self.collection = collection
        self.timeout = timeout
        self.last_search_hit_count = 0

    async def initialize(self, dimension: int) -> None:
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            response = await client.get(f"{self.url}/collections/{self.collection}")
            if response.status_code == 404:
                response = await client.put(
                    f"{self.url}/collections/{self.collection}",
                    json={"vectors": {"size": dimension, "distance": "Cosine"}},
                )
            response.raise_for_status()
            details = await client.get(f"{self.url}/collections/{self.collection}")
            details.raise_for_status()
            vectors = details.json()["result"]["config"]["params"]["vectors"]
            if not isinstance(vectors, dict) or int(vectors.get("size", 0)) != dimension:
                raise KnowledgeUnavailableError(
                    "Local Knowledge Store vector schema is incompatible with this release"
                )
            for field in self._indexed_fields:
                index = await client.put(
                    f"{self.url}/collections/{self.collection}/index?wait=true",
                    json={"field_name": field, "field_schema": "keyword"},
                )
                if index.status_code not in {200, 409}:
                    index.raise_for_status()

    async def health_check(self) -> bool:
        health = await self.health_status()
        return health.qdrant_reachable and health.collection_available

    async def health_status(self) -> KnowledgeStoreHealth:
        try:
            async with httpx.AsyncClient(timeout=self.timeout) as client:
                engine_version: str | None = None
                try:
                    root = await client.get(f"{self.url}/")
                    if root.is_success:
                        candidate = root.json().get("version")
                        if isinstance(candidate, str) and candidate.strip():
                            engine_version = candidate.strip()
                except (httpx.HTTPError, ValueError, AttributeError):
                    # Runtime version visibility is best-effort and does not determine health.
                    pass
                response = await client.get(f"{self.url}/healthz")
                if not response.is_success:
                    return KnowledgeStoreHealth(False, False, engine_version, False)
                collection = await client.get(f"{self.url}/collections/{self.collection}")
            collection_available = collection.is_success
            statistics_readable = False
            if collection_available:
                try:
                    result = collection.json().get("result")
                    statistics_readable = isinstance(result, dict)
                except (ValueError, AttributeError):
                    statistics_readable = False
            return KnowledgeStoreHealth(
                True, collection_available, engine_version, statistics_readable
            )
        except httpx.HTTPError:
            return KnowledgeStoreHealth(False, False)

    @staticmethod
    def _filter(tenant_id: UUID, filters: dict[str, str] | None = None) -> dict[str, Any]:
        must: list[dict[str, Any]] = [
            {"key": "tenant_id", "match": {"value": str(tenant_id)}},
            {"key": "schema_version", "match": {"value": KNOWLEDGE_SCHEMA_VERSION}},
        ]
        must.extend(
            {"key": key, "match": {"value": value}}
            for key, value in (filters or {}).items()
            if value
        )
        return {"must": must}

    async def upsert_chunks(self, tenant_id: UUID, chunks: list[KnowledgeChunk]) -> None:
        if not chunks:
            return
        points = []
        for chunk in chunks:
            payload = dict(chunk.metadata)
            payload.update(
                {
                    "tenant_id": str(tenant_id),
                    "document_id": str(chunk.document_id),
                    "chunk_id": str(chunk.chunk_id),
                    "content": chunk.content,
                    "schema_version": KNOWLEDGE_SCHEMA_VERSION,
                }
            )
            points.append({"id": str(chunk.chunk_id), "vector": chunk.vector, "payload": payload})
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            response = await client.put(
                f"{self.url}/collections/{self.collection}/points?wait=true",
                json={"points": points},
            )
        response.raise_for_status()

    async def search(
        self,
        tenant_id: UUID,
        vector: list[float],
        top_k: int,
        filters: dict[str, str] | None = None,
    ) -> list[KnowledgeSearchResult]:
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            response = await client.post(
                f"{self.url}/collections/{self.collection}/points/search",
                json={
                    "vector": vector,
                    "limit": top_k,
                    "with_payload": True,
                    "filter": self._filter(tenant_id, filters),
                },
            )
        response.raise_for_status()
        raw_hits = response.json().get("result") or []
        self.last_search_hit_count = len(raw_hits)
        results: list[KnowledgeSearchResult] = []
        for hit in raw_hits:
            payload = hit.get("payload") or {}
            try:
                results.append(
                    KnowledgeSearchResult(
                        document_id=UUID(str(payload["document_id"])),
                        chunk_id=UUID(str(payload["chunk_id"])),
                        content=str(payload["content"]),
                        score=float(hit["score"]),
                        source=str(payload.get("source") or "Documents"),
                        metadata={
                            key: value
                            for key, value in payload.items()
                            if key not in {"tenant_id", "document_id", "chunk_id", "content"}
                        },
                    )
                )
            except (KeyError, TypeError, ValueError):
                continue
        return results

    async def _delete(self, tenant_id: UUID, filters: dict[str, str] | None = None) -> None:
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            response = await client.post(
                f"{self.url}/collections/{self.collection}/points/delete?wait=true",
                json={"filter": self._filter(tenant_id, filters)},
            )
        response.raise_for_status()

    async def delete_document(self, tenant_id: UUID, document_id: UUID) -> None:
        await self._delete(tenant_id, {"document_id": str(document_id)})

    async def delete_tenant_knowledge(self, tenant_id: UUID) -> None:
        await self._delete(tenant_id)

    async def count(self, tenant_id: UUID) -> int:
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            response = await client.post(
                f"{self.url}/collections/{self.collection}/points/count",
                json={"filter": self._filter(tenant_id), "exact": True},
            )
        response.raise_for_status()
        return int(response.json()["result"]["count"])


class LocalKnowledgeService:
    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        store: KnowledgeStore | None = None,
        sanitizer: ContentSanitizer | None = None,
    ) -> None:
        self.session = session
        self.settings = settings
        self.store = store or QdrantKnowledgeStore(
            settings.qdrant_url, settings.qdrant_collection, settings.qdrant_timeout_seconds
        )
        self.sanitizer = sanitizer or ContentSanitizer()
        self.embeddings = LocalHashEmbeddingService(settings.knowledge_embedding_dimension)
        self.operations = SqlAlchemyOperationsRepository(session)

    async def tenant_id(self) -> UUID:
        product = await self.operations.get_settings()
        if not product.tenant_id:
            raise KnowledgeIdentityError(
                "Register the connector before indexing or searching local knowledge"
            )
        try:
            return UUID(product.tenant_id)
        except ValueError as exc:
            raise KnowledgeIdentityError(
                "The registered connector tenant identity is invalid"
            ) from exc

    async def initialize(self) -> None:
        try:
            await self.store.initialize(self.embeddings.dimension)
        except (httpx.HTTPError, KeyError, ValueError) as exc:
            raise KnowledgeUnavailableError("Local Knowledge Store initialization failed") from exc

    async def health_check(self) -> bool:
        return await self.store.health_check()

    async def health_status(self) -> KnowledgeStoreHealth:
        status = getattr(self.store, "health_status", None)
        if status is not None:
            return await status()
        healthy = await self.store.health_check()
        return KnowledgeStoreHealth(healthy, healthy)

    async def index_document(self, document: DocumentModel) -> int:
        tenant_id = await self.tenant_id()
        if document.owner_tenant_id and document.owner_tenant_id != str(tenant_id):
            raise KnowledgeIdentityError("Document tenant ownership does not match registration")
        path = self.settings.managed_documents_root / document.relative_path
        if not path.is_file() or path.is_symlink():
            raise FileNotFoundError("Document content is unavailable")
        sections = self._parse(path, document.mime_type)
        chunks = self._chunks(document, sections)
        try:
            await self.store.delete_document(tenant_id, document.id)
            await self.store.upsert_chunks(tenant_id, chunks)
        except (httpx.HTTPError, KeyError, ValueError) as exc:
            raise KnowledgeUnavailableError(
                "Local Knowledge Store is temporarily unavailable"
            ) from exc
        document.knowledge_status = "INDEXED"
        document.indexed_content_hash = document.content_hash
        document.indexed_chunk_count = len(chunks)
        document.knowledge_indexed_at = datetime.now(UTC)
        document.knowledge_error = None
        document.local_status = "INDEXED"
        document.delivery_status = "LOCAL_ONLY"
        await self.session.commit()
        logger.info(
            "Document indexed document_id=%s chunk_count=%s",
            str(document.id),
            len(chunks),
        )
        return len(chunks)

    async def delete_document(self, document_id: UUID) -> None:
        tenant_id = await self.tenant_id()
        try:
            await self.store.delete_document(tenant_id, document_id)
        except (httpx.HTTPError, KeyError, ValueError) as exc:
            raise KnowledgeUnavailableError(
                "Local Knowledge Store is temporarily unavailable"
            ) from exc
        document = await self.session.get(DocumentModel, document_id)
        if document is not None:
            document.knowledge_status = "DELETED"
            document.indexed_chunk_count = 0
            document.indexed_content_hash = None
            document.knowledge_error = None
            document.delivery_status = "LOCAL_ONLY"
            await self.session.commit()
        logger.info("Document removed from Local Knowledge Store document_id=%s", document_id)

    async def process_pending(self, limit: int = 10) -> int:
        documents = list(
            (
                await self.session.scalars(
                    select(DocumentModel)
                    .where(
                        DocumentModel.knowledge_status.in_(("PENDING", "FAILED", "DELETE_PENDING"))
                    )
                    .order_by(DocumentModel.updated_at)
                    .limit(limit)
                )
            ).all()
        )
        processed = 0
        for document in documents:
            try:
                if document.deletion_requested or document.knowledge_status == "DELETE_PENDING":
                    await self.delete_document(document.id)
                elif document.indexed_content_hash != document.content_hash:
                    await self.index_document(document)
                processed += 1
            except KnowledgeIdentityError:
                document.knowledge_status = "PENDING"
                document.knowledge_error = "Connector registration is required."
                await self.session.commit()
                break
            except Exception as exc:
                document.knowledge_status = "FAILED"
                document.knowledge_error = "Local knowledge indexing is temporarily unavailable."
                document.local_status = "INDEX_FAILED"
                await self.session.commit()
                logger.warning(
                    "Document indexing failed document_id=%s error_type=%s",
                    document.id,
                    type(exc).__name__,
                )
        return processed

    async def search(
        self, query: str, top_k: int, document_id: UUID | None = None
    ) -> list[KnowledgeSearchResult]:
        started = time.monotonic()
        tenant_id = await self.tenant_id()
        if document_id is not None:
            document = await self.session.get(DocumentModel, document_id)
            if (
                document is None
                or document.deletion_requested
                or (document.owner_tenant_id and document.owner_tenant_id != str(tenant_id))
            ):
                raise KnowledgeIdentityError("Document filter is not available for this tenant")
        sanitized_query = self.sanitizer.sanitize(query)
        vector = self.embeddings.embed([sanitized_query])[0]
        requested_limit = min(top_k, 20)
        candidate_limit = min(max(requested_limit * 4, 20), _SEARCH_CANDIDATE_LIMIT)
        logger.info(
            "knowledge search started query_length=%s tenant_id=%s collection=%s "
            "embedding_model=%s requested_result_limit=%s candidate_limit=%s "
            "qdrant_score_threshold=none evidence_score_threshold=%.2f document_id=%s",
            len(sanitized_query),
            tenant_id,
            getattr(self.store, "collection", self.settings.qdrant_collection),
            self.embeddings.model,
            requested_limit,
            candidate_limit,
            MINIMUM_EVIDENCE_SCORE,
            document_id,
        )
        try:
            candidates = await self.store.search(
                tenant_id,
                vector,
                candidate_limit,
                {"document_id": str(document_id)} if document_id else None,
            )
        except (httpx.HTTPError, KeyError, ValueError) as exc:
            raise KnowledgeUnavailableError(
                "Local Knowledge Store is temporarily unavailable"
            ) from exc
        accepted: list[KnowledgeSearchResult] = []
        for candidate in candidates:
            lexical_score, matched_terms = _lexical_relevance(
                sanitized_query, candidate.content
            )
            calibrated_score = (
                _LEXICAL_WEIGHT * lexical_score
                + _VECTOR_WEIGHT * max(0.0, min(1.0, candidate.score))
            )
            if (
                matched_terms < MINIMUM_LEXICAL_MATCHES
                or calibrated_score < MINIMUM_EVIDENCE_SCORE
            ):
                continue
            metadata = dict(candidate.metadata)
            metadata.update(
                {
                    "vector_score": candidate.score,
                    "lexical_score": lexical_score,
                    "matched_query_terms": matched_terms,
                    "retrieval_score_model": "local-hybrid-v1",
                }
            )
            accepted.append(
                KnowledgeSearchResult(
                    document_id=candidate.document_id,
                    chunk_id=candidate.chunk_id,
                    content=candidate.content,
                    score=calibrated_score,
                    source=candidate.source,
                    metadata=metadata,
                )
            )
        accepted.sort(key=lambda item: item.score, reverse=True)
        accepted = accepted[:requested_limit]
        product = await self.operations.get_settings()
        product.last_successful_knowledge_search_at = datetime.now(UTC)
        await self.session.commit()
        rejection_reason = (
            "qdrant_returned_zero_results"
            if not candidates
            else "qdrant_results_rejected_by_evidence_filter"
            if not accepted
            else "none"
        )
        top = accepted[0] if accepted else candidates[0] if candidates else None
        top_vector_score = (
            top.metadata.get("vector_score", top.score) if top is not None else None
        )
        qdrant_result_count = int(
            getattr(self.store, "last_search_hit_count", len(candidates))
        )
        logger.info(
            "knowledge search completed qdrant_result_count=%s parsed_result_count=%s "
            "accepted_result_count=%s top_score=%s top_vector_score=%s document_id=%s "
            "document_filename=%s chunk_index=%s "
            "rejection_reason=%s elapsed_ms=%.1f",
            qdrant_result_count,
            len(candidates),
            len(accepted),
            round(top.score, 6) if top else None,
            round(float(top_vector_score), 6) if top_vector_score is not None else None,
            top.document_id if top else None,
            top.metadata.get("filename") if top else None,
            top.metadata.get("chunk_index") if top else None,
            rejection_reason,
            (time.monotonic() - started) * 1000,
        )
        return accepted

    async def stats(self) -> KnowledgeStats:
        indexed_filter = (
            DocumentModel.knowledge_status == "INDEXED",
            DocumentModel.deletion_requested.is_(False),
        )
        documents = int(
            await self.session.scalar(select(func.count(DocumentModel.id)).where(*indexed_filter))
            or 0
        )
        chunk_count = int(
            await self.session.scalar(
                select(func.coalesce(func.sum(DocumentModel.indexed_chunk_count), 0)).where(
                    *indexed_filter
                )
            )
            or 0
        )
        pending_count = int(
            await self.session.scalar(
                select(func.count(DocumentModel.id)).where(
                    DocumentModel.knowledge_status == "PENDING",
                    DocumentModel.deletion_requested.is_(False),
                )
            )
            or 0
        )
        failed_count = int(
            await self.session.scalar(
                select(func.count(DocumentModel.id)).where(
                    DocumentModel.knowledge_status == "FAILED",
                    DocumentModel.deletion_requested.is_(False),
                )
            )
            or 0
        )
        last_activity = await self.session.scalar(
            select(func.max(DocumentModel.knowledge_indexed_at)).where(*indexed_filter)
        )
        product = await self.operations.get_settings()
        health = await self.health_status()
        return KnowledgeStats(
            documents,
            chunk_count,
            pending_count,
            failed_count,
            last_activity,
            health.qdrant_reachable,
            health.collection_available,
            documents > 0,
            chunk_count > 0,
            health.qdrant_reachable
            and health.collection_available
            and health.statistics_readable,
            True if product.last_successful_knowledge_search_at is not None else None,
            product.last_successful_knowledge_search_at,
            health.engine_version,
            health.statistics_readable,
        )

    def _chunks(
        self, document: DocumentModel, sections: list[KnowledgeSection]
    ) -> list[KnowledgeChunk]:
        texts: list[tuple[str, KnowledgeSection]] = []
        maximum = self.settings.knowledge_max_chunk_characters
        overlap = min(400, maximum // 10)
        for section in sections:
            sanitized = self.sanitizer.sanitize(section.text)
            start = 0
            while sanitized and start < len(sanitized):
                end = min(start + maximum, len(sanitized))
                if end < len(sanitized):
                    boundary = sanitized.rfind(" ", start, end)
                    if boundary > start + maximum // 2:
                        end = boundary
                content = sanitized[start:end].strip()
                if content:
                    texts.append((content, section))
                if end >= len(sanitized):
                    break
                start = max(start + 1, end - overlap)
        vectors = self.embeddings.embed([text for text, _section in texts])
        chunks: list[KnowledgeChunk] = []
        for index, ((content, section), vector) in enumerate(zip(texts, vectors, strict=True)):
            chunk_id = uuid5(document.id, f"{document.content_hash}:{index}")
            chunks.append(
                KnowledgeChunk(
                    chunk_id=chunk_id,
                    document_id=document.id,
                    content=content,
                    vector=vector,
                    metadata={
                        "source_id": str(document.source_id),
                        "source": "Documents",
                        "filename": document.filename,
                        "content_hash": document.content_hash,
                        "version_id": str(uuid5(document.id, document.content_hash)),
                        "chunk_index": index,
                        "page_number": section.page_number,
                        "sheet_name": section.sheet_name,
                        "row_start": section.row_start,
                        "row_end": section.row_end,
                        "section_title": section.section_title,
                        "embedding_model": self.embeddings.model,
                    },
                )
            )
        return chunks

    @staticmethod
    def _parse(path: Path, mime_type: str) -> list[KnowledgeSection]:
        extension = path.suffix.casefold()
        with path.open("rb") as stream:
            if extension in {".txt", ".md"}:
                return [KnowledgeSection(stream.read().decode("utf-8", errors="replace"))]
            if extension == ".csv":
                text = stream.read().decode("utf-8-sig", errors="replace")
                rows = list(csv.reader(io.StringIO(text)))
                header = rows[0] if rows else []
                return [
                    KnowledgeSection(
                        "\n".join(
                            f"{header[column]}: {value}" if column < len(header) else value
                            for column, value in enumerate(row)
                        ),
                        row_start=index,
                        row_end=index,
                    )
                    for index, row in enumerate(rows[1:], start=2)
                ]
            if extension == ".pdf":
                return LocalKnowledgeService._parse_pdf(stream)
            if extension == ".docx":
                return LocalKnowledgeService._parse_docx(stream)
            if extension == ".xlsx":
                return LocalKnowledgeService._parse_xlsx(stream)
        raise ValueError(f"Unsupported document type: {mime_type}")

    @staticmethod
    def _parse_pdf(stream: BinaryIO) -> list[KnowledgeSection]:
        from pypdf import PdfReader  # type: ignore[import-not-found]

        reader = PdfReader(stream)
        if reader.is_encrypted:
            raise ValueError("Encrypted PDF documents are unsupported")
        return [
            KnowledgeSection(page.extract_text() or "", page_number=index)
            for index, page in enumerate(reader.pages, start=1)
        ]

    @staticmethod
    def _parse_docx(stream: BinaryIO) -> list[KnowledgeSection]:
        from docx import Document  # type: ignore[import-not-found]

        document = Document(stream)
        return [
            KnowledgeSection(paragraph.text)
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        ]

    @staticmethod
    def _parse_xlsx(stream: BinaryIO) -> list[KnowledgeSection]:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        workbook = load_workbook(stream, read_only=True, data_only=True)
        return [
            KnowledgeSection(
                " | ".join("" if value is None else str(value) for value in row),
                sheet_name=sheet.title,
                row_start=row_number,
                row_end=row_number,
            )
            for sheet in workbook.worksheets
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1)
            if any(value is not None and str(value) for value in row)
        ]
